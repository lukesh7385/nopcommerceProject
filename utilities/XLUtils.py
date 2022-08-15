import inspect
import logging
import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columnno).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum, columnno).fill = redFill
    workbook.save(file)


# def custom_logger():
#     # set class/method name from where its called
#     logger_name = inspect.stack()[1][3]
#     # create logger
#     logger = logging.getLogger(logger_name)
#     logger.setLevel(logging.INFO)
#     # create console handler or file handler and set the log level
#     fh = logging.FileHandler(filename='.\\Logs\\automation.log')
#     # create formatter - how you want our logs to be formatted
#     formatter = logging.Formatter('%(asctime)s: %(levelname)s: %(massage)s', datefmt='%m/%d/%Y %I:%M:%S %p')
#     # add formatter to console or file handler
#     fh.setFormatter(formatter)
#     # add console handler to logger
#     logger.addHandler(fh)
#     return logger


